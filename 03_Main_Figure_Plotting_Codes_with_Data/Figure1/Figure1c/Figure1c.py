#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Figure 1c for PulseBat-combo: phi-SOH scatter plots at SOC 25%, 50%, and 75%.
"""

from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import FormatStrFormatter
from openpyxl import load_workbook


ROOT = Path(r"E:\Datasets\PulseBat_combo\Figure1\Figure1c")
FEATURE_ROOT = Path(r"E:\Datasets\PulseBat\PulseBat_Fig\Figure2\Fts-For-Model")

FIG_W_CM = 8.5
FIG_H_CM = 6.51

SOC_SHEETS = [
    ("SOC25", "25%"),
    ("SOC50", "50%"),
    ("SOC75", "75%"),
]

CRATES = [
    ("0.5C", "Hyst_M3_0.5C"),
    ("1.0C", "Hyst_M3_1C"),
    ("1.5C", "Hyst_M3_1.5C"),
    ("2.0C", "Hyst_M3_2C"),
    ("2.5C", "Hyst_M3_2.5C"),
]

# Keep these colors aligned with Figure1a.py default_groups().
CAPACITIES = [
    ("20Ah", "20 Ah", "20Ah LFP", FEATURE_ROOT / "20Ah LFP" / "LFP_20Ah_W_5000.xlsx", "#077FFF"),
    ("35Ah", "35 Ah", "35Ah LFP", FEATURE_ROOT / "35Ah LFP" / "LFP_35Ah_W_5000.xlsx", "#FF972F"),
    ("68Ah", "68 Ah", "68Ah LFP", FEATURE_ROOT / "68Ah LFP" / "LFP_68Ah_W_5000.xlsx", "#50D541"),
]

POINT_SIZE = 5.4
POINT_ALPHA = 0.48
POINT_EDGE_COLOR = "0.55"
POINT_EDGE_LW = 0.1
FIT_LINE_LW = 0.9
FIT_LINE_ALPHA = 0.9
FIT_LINE_COLOR = "#12355B"


@dataclass(frozen=True)
class Record:
    soc_sheet: str
    soc_label: str
    capacity_key: str
    capacity_label: str
    group: str
    color: str
    soh_pct: float
    phi_mv: Dict[str, float]


@dataclass(frozen=True)
class FitResult:
    soc_sheet: str
    soc_label: str
    crate_label: str
    capacity_key: str
    capacity_label: str
    n: int
    slope: float
    intercept: float
    r2: float
    x_min: float
    x_max: float


def cm_to_in(cm: float) -> float:
    return cm / 2.54


def setup_style() -> None:
    mpl.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Arial", "DejaVu Sans"],
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "axes.linewidth": 0.75,
        "axes.labelsize": 8,
        "xtick.labelsize": 6,
        "ytick.labelsize": 6,
    })


def is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and math.isfinite(float(value))


def load_sheet_records(
    path: Path,
    soc_sheet: str,
    soc_label: str,
    capacity_key: str,
    capacity_label: str,
    group: str,
    color: str,
) -> List[Record]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if soc_sheet not in wb.sheetnames:
            raise RuntimeError(f"{path} has no sheet {soc_sheet}")
        ws = wb[soc_sheet]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        col_idx = {str(name): i for i, name in enumerate(headers) if name is not None}
        required = ["SOH"] + [col for _, col in CRATES]
        missing = [col for col in required if col not in col_idx]
        if missing:
            raise RuntimeError(f"{path} {soc_sheet} missing columns: {missing}")

        records: List[Record] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            soh = row[col_idx["SOH"]]
            if not is_number(soh):
                continue

            phi_mv: Dict[str, float] = {}
            ok = True
            for crate_label, col in CRATES:
                value = row[col_idx[col]]
                if not is_number(value):
                    ok = False
                    break
                phi_mv[crate_label] = float(value) * 1000.0
            if not ok:
                continue

            records.append(Record(
                soc_sheet=soc_sheet,
                soc_label=soc_label,
                capacity_key=capacity_key,
                capacity_label=capacity_label,
                group=group,
                color=color,
                soh_pct=float(soh) * 100.0,
                phi_mv=phi_mv,
            ))
        return records
    finally:
        wb.close()


def load_records() -> List[Record]:
    records: List[Record] = []
    for soc_sheet, soc_label in SOC_SHEETS:
        for capacity_key, capacity_label, group, path, color in CAPACITIES:
            records.extend(load_sheet_records(
                path=path,
                soc_sheet=soc_sheet,
                soc_label=soc_label,
                capacity_key=capacity_key,
                capacity_label=capacity_label,
                group=group,
                color=color,
            ))
    return records


def subset_records(
    records: Sequence[Record],
    soc_sheet: str,
    capacity_key: str | None = None,
) -> List[Record]:
    out = [r for r in records if r.soc_sheet == soc_sheet]
    if capacity_key is not None:
        out = [r for r in out if r.capacity_key == capacity_key]
    return out


def padded_limits(values: Iterable[float], min_pad: float, frac: float = 0.08) -> Tuple[float, float]:
    arr = np.asarray(list(values), dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        raise RuntimeError("No finite values for axis limits.")
    vmin = float(np.nanmin(arr))
    vmax = float(np.nanmax(arr))
    if abs(vmax - vmin) < 1e-12:
        return vmin - min_pad, vmax + min_pad
    pad = max(min_pad, frac * (vmax - vmin))
    return vmin - pad, vmax + pad


def three_ticks(vmin: float, vmax: float) -> List[float]:
    return [vmin, 0.5 * (vmin + vmax), vmax]


def padded_integer_limits(values: Iterable[float], pad: float = 1.5) -> Tuple[float, float]:
    arr = np.asarray(list(values), dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        raise RuntimeError("No finite values for axis limits.")
    return float(math.floor(float(np.nanmin(arr)) - pad)), float(math.ceil(float(np.nanmax(arr)) + pad))


def linear_fit(x: np.ndarray, y: np.ndarray) -> Tuple[float, float, float]:
    mask = np.isfinite(x) & np.isfinite(y)
    x = x[mask]
    y = y[mask]
    if x.size < 2 or np.nanstd(x) <= 0:
        return float("nan"), float("nan"), float("nan")

    slope, intercept = np.polyfit(x, y, deg=1)
    y_hat = slope * x + intercept
    ss_res = float(np.sum((y - y_hat) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = float("nan") if ss_tot <= 0 else 1.0 - ss_res / ss_tot
    return float(slope), float(intercept), float(r2)


def fit_line_segment(
    slope: float,
    intercept: float,
    data_x_min: float,
    data_x_max: float,
    axis_x_min: float,
    axis_x_max: float,
    axis_y_min: float,
    axis_y_max: float,
    extend_frac: float = 0.35,
) -> Tuple[np.ndarray, np.ndarray]:
    data_span = max(1e-12, data_x_max - data_x_min)
    x_min = max(axis_x_min, data_x_min - extend_frac * data_span)
    x_max = min(axis_x_max, data_x_max + extend_frac * data_span)
    x_line = np.linspace(x_min, x_max, 160)
    y_line = slope * x_line + intercept
    mask = np.isfinite(y_line) & (y_line >= axis_y_min) & (y_line <= axis_y_max)
    return x_line[mask], y_line[mask]


def write_fit_csv(fits: Sequence[FitResult], out_dir: Path) -> None:
    path = out_dir / "Figure1c_linear_fits.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "soc_sheet",
            "soc_label",
            "crate",
            "capacity",
            "capacity_label",
            "n",
            "slope_soh_pct_per_mv",
            "intercept_soh_pct",
            "r2",
            "x_min_mv",
            "x_max_mv",
        ])
        for fit in fits:
            writer.writerow([
                fit.soc_sheet,
                fit.soc_label,
                fit.crate_label,
                fit.capacity_key,
                fit.capacity_label,
                fit.n,
                fit.slope,
                fit.intercept,
                fit.r2,
                fit.x_min,
                fit.x_max,
            ])


def write_points_csv(records: Sequence[Record], out_dir: Path) -> None:
    path = out_dir / "Figure1c_phi_soh_points.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["soc_sheet", "soc_label", "capacity", "capacity_label", "SOH_percent", *[c for c, _ in CRATES]])
        for r in records:
            writer.writerow([
                r.soc_sheet,
                r.soc_label,
                r.capacity_key,
                r.capacity_label,
                r.soh_pct,
                *[r.phi_mv[crate_label] for crate_label, _ in CRATES],
            ])


def plot(records: Sequence[Record], out_dir: Path, fig_w_cm: float, fig_h_cm: float) -> None:
    setup_style()
    out_dir.mkdir(parents=True, exist_ok=True)

    fig, axes = plt.subplots(
        len(SOC_SHEETS),
        len(CRATES),
        figsize=(cm_to_in(fig_w_cm), cm_to_in(fig_h_cm)),
        sharey=False,
    )
    fig.subplots_adjust(left=0.11, right=0.965, bottom=0.17, top=0.955, wspace=0.30, hspace=0.62)

    fits: List[FitResult] = []
    for row_idx, (soc_sheet, soc_label) in enumerate(SOC_SHEETS):
        soc_records = subset_records(records, soc_sheet)
        y_min, y_max = padded_integer_limits((r.soh_pct for r in soc_records), pad=1.5)
        y_ticks = three_ticks(y_min, y_max)
        for col_idx, (crate_label, _) in enumerate(CRATES):
            ax = axes[row_idx, col_idx]
            x_min, x_max = padded_limits((r.phi_mv[crate_label] for r in soc_records), min_pad=0.45, frac=0.14)

            for capacity_key, capacity_label, _, _, color in CAPACITIES:
                cap_records = subset_records(records, soc_sheet, capacity_key)
                x = np.asarray([r.phi_mv[crate_label] for r in cap_records], dtype=float)
                y = np.asarray([r.soh_pct for r in cap_records], dtype=float)
                slope, intercept, r2 = linear_fit(x, y)
                finite = np.isfinite(x) & np.isfinite(y)

                if np.isfinite(slope) and np.isfinite(intercept) and np.any(finite):
                    x_fit, y_fit = fit_line_segment(
                        slope=slope,
                        intercept=intercept,
                        data_x_min=float(np.nanmin(x[finite])),
                        data_x_max=float(np.nanmax(x[finite])),
                        axis_x_min=x_min,
                        axis_x_max=x_max,
                        axis_y_min=y_min,
                        axis_y_max=y_max,
                    )
                    if x_fit.size >= 2:
                        ax.plot(
                            x_fit,
                            y_fit,
                            color=FIT_LINE_COLOR,
                            lw=FIT_LINE_LW,
                            alpha=FIT_LINE_ALPHA,
                            linestyle="--",
                            zorder=1,
                        )

                ax.scatter(
                    x,
                    y,
                    s=POINT_SIZE,
                    color=color,
                    alpha=POINT_ALPHA,
                    edgecolors=POINT_EDGE_COLOR,
                    linewidths=POINT_EDGE_LW,
                    rasterized=True,
                    zorder=2,
                )

                fits.append(FitResult(
                    soc_sheet=soc_sheet,
                    soc_label=soc_label,
                    crate_label=crate_label,
                    capacity_key=capacity_key,
                    capacity_label=capacity_label,
                    n=int(np.sum(finite)),
                    slope=slope,
                    intercept=intercept,
                    r2=r2,
                    x_min=float(np.nanmin(x[finite])) if np.any(finite) else float("nan"),
                    x_max=float(np.nanmax(x[finite])) if np.any(finite) else float("nan"),
                ))

            ax.set_xlim(x_min, x_max)
            ax.set_ylim(y_min, y_max)
            ax.set_xticks(three_ticks(x_min, x_max))
            ax.set_yticks(y_ticks)
            ax.xaxis.set_major_formatter(FormatStrFormatter("%.1f"))
            ax.yaxis.set_major_formatter(FormatStrFormatter("%.0f"))
            ax.tick_params(length=2.0, width=0.6, pad=1.0)

            if col_idx != 0:
                ax.tick_params(labelleft=False)
            for label in ax.get_xticklabels():
                label.set_rotation(45)
                label.set_ha("right")
                label.set_rotation_mode("anchor")

            for spine in ax.spines.values():
                spine.set_linewidth(0.75)

    fig.text(0.5, 0.045, r"$\phi$ (mV)", ha="center", va="center", fontsize=8)
    fig.text(0.028, 0.56, "State of health (%)", ha="center", va="center", rotation=90, fontsize=8)

    write_fit_csv(fits, out_dir)
    write_points_csv(records, out_dir)

    png_path = out_dir / "Figure1c.png"
    pdf_path = out_dir / "Figure1c.pdf"
    fig.savefig(png_path, dpi=600)
    fig.savefig(pdf_path)
    plt.close(fig)

    print(f"[OK] Saved: {png_path}")
    print(f"[OK] Saved: {pdf_path}")
    print(f"[OK] Saved: {out_dir / 'Figure1c_linear_fits.csv'}")
    print(f"[OK] Saved: {out_dir / 'Figure1c_phi_soh_points.csv'}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Plot Figure1c phi-SOH scatter grids.")
    parser.add_argument("--out_dir", type=str, default=str(ROOT))
    parser.add_argument("--fig_w_cm", type=float, default=FIG_W_CM)
    parser.add_argument("--fig_h_cm", type=float, default=FIG_H_CM)
    args = parser.parse_args()

    records = load_records()
    print(f"[INFO] Loaded records: {len(records)}")
    plot(records, Path(args.out_dir), float(args.fig_w_cm), float(args.fig_h_cm))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
