#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Sequence

import matplotlib

matplotlib.use("Agg")
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.cm import ScalarMappable
from matplotlib.colors import Normalize
from openpyxl import load_workbook


ROOT = Path(r"E:\Datasets\PulseBat_combo\Figure1\Figure1d")
FEATURE_ROOT = Path(r"E:\Datasets\PulseBat\PulseBat_Fig\Figure2\Fts-For-Model")

FIG_W_CM = 5.0
FIG_H_CM = 6.51

SOC_SHEETS = [
    ("SOC25", "25% State of charge"),
    ("SOC50", "50% State of charge"),
    ("SOC75", "75% State of charge"),
]

CAPACITIES = [
    ("20Ah", "20 Ah", FEATURE_ROOT / "20Ah LFP" / "LFP_20Ah_W_5000.xlsx"),
    ("35Ah", "35 Ah", FEATURE_ROOT / "35Ah LFP" / "LFP_35Ah_W_5000.xlsx"),
    ("68Ah", "68 Ah", FEATURE_ROOT / "68Ah LFP" / "LFP_68Ah_W_5000.xlsx"),
]

CRATES = [
    ("0.5C", "Hyst_M3_0.5C"),
    ("1.0C", "Hyst_M3_1C"),
    ("1.5C", "Hyst_M3_1.5C"),
    ("2.0C", "Hyst_M3_2C"),
    ("2.5C", "Hyst_M3_2.5C"),
]

GRID_EDGE_COLOR = "#0A2F42"
GRID_EDGE_LW = 0.65
SQUARE_EDGE_COLOR = "#777777"
SQUARE_EDGE_LW = 0.25
MIN_SQUARE_FRAC = 0.18
MAX_SQUARE_FRAC = 0.86


@dataclass(frozen=True)
class CorrelationRecord:
    soc_sheet: str
    soc_label: str
    capacity: str
    capacity_label: str
    crate: str
    n: int
    spearman_rho: float


def cm_to_in(cm: float) -> float:
    return cm / 2.54


def setup_style() -> None:
    mpl.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Arial", "DejaVu Sans"],
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "axes.linewidth": 0.75,
        "xtick.labelsize": 7,
        "ytick.labelsize": 7,
    })


def is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and math.isfinite(float(value))


def rankdata(values: np.ndarray) -> np.ndarray:
    order = np.argsort(values, kind="mergesort")
    ranks = np.empty(len(values), dtype=float)
    sorted_values = values[order]
    i = 0
    while i < len(values):
        j = i + 1
        while j < len(values) and sorted_values[j] == sorted_values[i]:
            j += 1
        avg_rank = 0.5 * (i + j - 1) + 1.0
        ranks[order[i:j]] = avg_rank
        i = j
    return ranks


def spearman_rho(x: np.ndarray, y: np.ndarray) -> float:
    mask = np.isfinite(x) & np.isfinite(y)
    x = x[mask]
    y = y[mask]
    if x.size < 3:
        return float("nan")
    rx = rankdata(x)
    ry = rankdata(y)
    if np.nanstd(rx) == 0 or np.nanstd(ry) == 0:
        return float("nan")
    return float(np.corrcoef(rx, ry)[0, 1])


def load_sheet_values(path: Path, sheet_name: str) -> Dict[str, List[float]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"{path} has no sheet {sheet_name}")
        ws = wb[sheet_name]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        col_idx = {str(name): i for i, name in enumerate(headers) if name is not None}
        required = ["SOH"] + [col for _, col in CRATES]
        missing = [col for col in required if col not in col_idx]
        if missing:
            raise RuntimeError(f"{path} {sheet_name} missing columns: {missing}")

        values: Dict[str, List[float]] = {"SOH": []}
        for crate_label, _ in CRATES:
            values[crate_label] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            soh = row[col_idx["SOH"]]
            if not is_number(soh):
                continue
            row_values = {}
            ok = True
            for crate_label, col in CRATES:
                phi = row[col_idx[col]]
                if not is_number(phi):
                    ok = False
                    break
                row_values[crate_label] = float(phi) * 1000.0
            if not ok:
                continue

            values["SOH"].append(float(soh) * 100.0)
            for crate_label, _ in CRATES:
                values[crate_label].append(row_values[crate_label])
        return values
    finally:
        wb.close()


def compute_correlations() -> List[CorrelationRecord]:
    out: List[CorrelationRecord] = []
    for soc_sheet, soc_label in SOC_SHEETS:
        for capacity, capacity_label, path in CAPACITIES:
            values = load_sheet_values(path, soc_sheet)
            y = np.asarray(values["SOH"], dtype=float)
            for crate_label, _ in CRATES:
                x = np.asarray(values[crate_label], dtype=float)
                rho = spearman_rho(x, y)
                out.append(CorrelationRecord(
                    soc_sheet=soc_sheet,
                    soc_label=soc_label,
                    capacity=capacity,
                    capacity_label=capacity_label,
                    crate=crate_label,
                    n=int(np.sum(np.isfinite(x) & np.isfinite(y))),
                    spearman_rho=rho,
                ))
    return out


def write_correlation_csv(records: Sequence[CorrelationRecord], out_dir: Path) -> None:
    path = out_dir / "Figure1d_spearman_rho.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["soc_sheet", "soc_label", "capacity", "capacity_label", "crate", "n", "spearman_rho"])
        for r in records:
            writer.writerow([r.soc_sheet, r.soc_label, r.capacity, r.capacity_label, r.crate, r.n, r.spearman_rho])


def rho_matrix(records: Sequence[CorrelationRecord], soc_sheet: str) -> np.ndarray:
    mat = np.full((len(CAPACITIES), len(CRATES)), np.nan, dtype=float)
    for i, (capacity, _, _) in enumerate(CAPACITIES):
        for j, (crate, _) in enumerate(CRATES):
            matched = [
                r for r in records
                if r.soc_sheet == soc_sheet and r.capacity == capacity and r.crate == crate
            ]
            if matched:
                mat[i, j] = matched[0].spearman_rho
    return mat


def square_size(rho: float) -> float:
    if not np.isfinite(rho):
        return 0.0
    strength = min(1.0, abs(float(rho)))
    return MIN_SQUARE_FRAC + (MAX_SQUARE_FRAC - MIN_SQUARE_FRAC) * strength


def plot(records: Sequence[CorrelationRecord], out_dir: Path, fig_w_cm: float, fig_h_cm: float) -> None:
    setup_style()
    out_dir.mkdir(parents=True, exist_ok=True)

    cmap = mpl.colormaps["Spectral"]

    fig = plt.figure(figsize=(cm_to_in(fig_w_cm), cm_to_in(fig_h_cm)))
    gs = fig.add_gridspec(
        len(SOC_SHEETS) + 1,
        1,
        height_ratios=[1.0, 1.0, 1.0, 0.12],
        hspace=0.42,
    )
    axes = [fig.add_subplot(gs[i, 0]) for i in range(len(SOC_SHEETS))]
    cax = fig.add_subplot(gs[-1, 0])
    fig.subplots_adjust(left=0.08, right=0.94, bottom=0.16, top=0.94)

    all_rhos = np.array([rec.spearman_rho for rec in records if np.isfinite(rec.spearman_rho)], dtype=float)
    vmin = float(np.nanmin(all_rhos))
    vmax = float(np.nanmax(all_rhos))
    norm = Normalize(vmin=vmin, vmax=vmax)

    for panel_idx, (ax, (soc_sheet, soc_label)) in enumerate(zip(axes, SOC_SHEETS)):
        mat = rho_matrix(records, soc_sheet)

        for i in range(len(CAPACITIES)):
            for j in range(len(CRATES)):
                rho = mat[i, j]
                size = square_size(rho)
                x0 = j + 0.5 - size / 2.0
                y0 = i + 0.5 - size / 2.0
                rect = plt.Rectangle(
                    (x0, y0),
                    size,
                    size,
                    facecolor=cmap(norm(rho)) if np.isfinite(rho) else "white",
                    edgecolor=SQUARE_EDGE_COLOR,
                    linewidth=SQUARE_EDGE_LW,
                )
                ax.add_patch(rect)

        for x in range(len(CRATES) + 1):
            ax.plot([x, x], [0, len(CAPACITIES)], color=GRID_EDGE_COLOR, lw=GRID_EDGE_LW)
        for y in range(len(CAPACITIES) + 1):
            ax.plot([0, len(CRATES)], [y, y], color=GRID_EDGE_COLOR, lw=GRID_EDGE_LW)

        ax.set_xlim(0, len(CRATES))
        ax.set_ylim(len(CAPACITIES), 0)
        ax.set_aspect("auto")
        ax.set_title("", fontsize=8, pad=2)
        ax.set_xticks(np.arange(len(CRATES)) + 0.5)
        if panel_idx == len(SOC_SHEETS) - 1:
            ax.set_xticklabels([label for label, _ in CRATES], fontsize=6)
        else:
            ax.set_xticklabels([])
        ax.set_yticks(np.arange(len(CAPACITIES)) + 0.5)
        ax.set_yticklabels(["" for _ in CAPACITIES], fontsize=6)
        ax.tick_params(length=0, pad=2)
        for spine in ax.spines.values():
            spine.set_visible(False)

    sm = ScalarMappable(norm=norm, cmap=cmap)
    sm.set_array([])
    cbar = fig.colorbar(sm, cax=cax, orientation="horizontal")
    cbar_ticks = np.linspace(vmin, vmax, 5)
    cbar.set_ticks(cbar_ticks)
    cbar.ax.set_xticklabels([f"{t:.2f}" for t in cbar_ticks], fontsize=6)
    cbar.set_label(r"$\rho$", fontsize=8, labelpad=2)
    cbar.ax.tick_params(length=2.0, width=0.6, pad=1.2)
    cbar.outline.set_linewidth(0.6)

    write_correlation_csv(records, out_dir)
    png_path = out_dir / "Figure1d.png"
    pdf_path = out_dir / "Figure1d.pdf"
    fig.savefig(png_path, dpi=600)
    fig.savefig(pdf_path)
    plt.close(fig)

    print(f"[OK] Saved: {png_path}")
    print(f"[OK] Saved: {pdf_path}")
    print(f"[OK] Saved: {out_dir / 'Figure1d_spearman_rho.csv'}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Plot Figure1d SOC-wise Spearman rho square heatmaps.")
    parser.add_argument("--out_dir", type=str, default=str(ROOT))
    parser.add_argument("--fig_w_cm", type=float, default=FIG_W_CM)
    parser.add_argument("--fig_h_cm", type=float, default=FIG_H_CM)
    args = parser.parse_args()

    records = compute_correlations()
    plot(records, Path(args.out_dir), float(args.fig_w_cm), float(args.fig_h_cm))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
